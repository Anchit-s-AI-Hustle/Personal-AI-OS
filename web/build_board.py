#!/usr/bin/env python3
"""
Generate web/data.json for the Kanban task board.

Source priority:
  1. SQLite DB (extracted_tasks)  — live source of truth, if present
  2. tasks.xlsx "All Tasks" tab    — the Sheets mirror fallback

Usage (run from repo root or anywhere):
    python web/build_board.py

The deployed board ships with whatever data.json is committed; it also lets you
drag-and-drop an .xlsx/.csv/.json export in the browser (nothing is uploaded).
"""
from __future__ import annotations
import json, sqlite3
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
OUT = Path(__file__).resolve().parent / "data.json"

# canonical task shape the board consumes
def blank():
    return {"id":"","source_type":"","task":"","task_description":"","rationale":"",
            "growth_pillar":"","deadline":"","urgency":"Medium","spoc":"","summary":"",
            "status":"open","done":False,"source_link":"","created_at":""}

def norm_status(status, done):
    d = str(done).strip().lower() in ("true","1","yes","done","y")
    s = str(status or "").strip().lower()
    if d or s in ("done","closed","complete","completed"): return "done"
    if s in ("dropped","cancelled","canceled","wontfix","won't do"): return "dropped"
    if s in ("in_progress","in progress","doing","wip","started"): return "in_progress"
    return "open"

def norm_urgency(p):
    p = str(p or "Medium").strip().title()
    return p if p in ("Low","Medium","High","Critical") else "Medium"

def from_sqlite():
    cands = [ROOT/"data"/"personal_ai_os.db", ROOT/"data"/"app.db", ROOT/"personal_ai_os.db"]
    try:
        from config import settings  # type: ignore
        p = getattr(settings, "db_path", None)
        if p: cands.insert(0, Path(p))
    except Exception:
        pass
    for db in cands:
        if not (db and Path(db).exists()): continue
        con = sqlite3.connect(str(db)); con.row_factory = sqlite3.Row
        try:
            rs = con.execute("SELECT id,source_type,task,task_description,rationale,"
                "growth_pillar,deadline,urgency,sender_or_speaker,summary,status,created_at "
                "FROM extracted_tasks ORDER BY created_at DESC").fetchall()
        except Exception:
            con.close(); continue
        con.close()
        out=[]
        for r in rs:
            t=blank()
            t.update(id=r["id"], source_type=r["source_type"] or "", task=r["task"] or "",
                task_description=r["task_description"] or "", rationale=r["rationale"] or "",
                growth_pillar=r["growth_pillar"] or "", deadline=r["deadline"] or "",
                urgency=norm_urgency(r["urgency"]), spoc=r["sender_or_speaker"] or "",
                summary=r["summary"] or "", created_at=str(r["created_at"] or ""))
            t["status"]=norm_status(r["status"], False); t["done"]=(t["status"]=="done")
            out.append(t)
        if out: return out, "sqlite"
    return None

def from_xlsx():
    try:
        import openpyxl
    except Exception:
        return [], "none"
    xp = ROOT/"tasks.xlsx"
    if not xp.exists(): return [], "none"
    wb = openpyxl.load_workbook(str(xp), read_only=True, data_only=True)
    sn = "All Tasks" if "All Tasks" in wb.sheetnames else (
         "Master Task List" if "Master Task List" in wb.sheetnames else wb.sheetnames[0])
    ws = wb[sn]
    rows = [r for r in ws.iter_rows(values_only=True) if r and any(c not in (None,"") for c in r)]
    if len(rows) < 2: return [], "none"
    hdr = [str(c).strip() if c else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    def g(r, *names):
        for n in names:
            if n in idx and idx[n] < len(r) and r[idx[n]] not in (None, ""):
                return r[idx[n]]
        return ""
    out=[]
    for i, r in enumerate(rows[1:], 1):
        t=blank()
        t.update(
            id=g(r,"_task_id") or i,
            source_type=g(r,"Source"),
            source_link=g(r,"Source Link"),
            task=g(r,"Task Heading","Task"),
            task_description=g(r,"Task Description"),
            rationale=g(r,"Why We're Doing This","Why We’re Doing This"),
            growth_pillar=g(r,"Growth Pillar"),
            deadline=g(r,"Task Deadline","Deadline"),
            urgency=norm_urgency(g(r,"Priority")),
            spoc=g(r,"SPOC"),
            summary=g(r,"Remarks") or g(r,"All Updates"),
            created_at=str(g(r,"Task Given On") or ""),
        )
        t["status"]=norm_status(g(r,"Status"), g(r,"Done?"))
        t["done"]=(t["status"]=="done")
        # collapse "Email from X" → "Email" for cleaner filter chips, keep person in spoc
        st=str(t["source_type"])
        if st.lower().startswith("email"):
            if not t["spoc"] and " from " in st: t["spoc"]=st.split(" from ",1)[1]
            t["source_type"]="Email"
        elif st.lower().startswith("google chat"): t["source_type"]="Google Chat"
        elif "voice" in st.lower() or "memo" in st.lower(): t["source_type"]="Voice memo"
        out.append(t)
    return out, "xlsx (All Tasks)"

def main():
    res = from_sqlite()
    if res: tasks, src = res
    else:   tasks, src = from_xlsx()
    for t in tasks:
        for k,v in list(t.items()):
            if v is None: t[k]=""
    payload={"generated_from":src,"count":len(tasks),"tasks":tasks}
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    done=sum(1 for t in tasks if t["status"]=="done")
    print(f"WROTE {OUT}  source={src}  tasks={len(tasks)}  done={done}")

if __name__ == "__main__":
    main()
