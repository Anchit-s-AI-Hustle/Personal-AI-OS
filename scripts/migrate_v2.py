"""
v2 schema migration — bring DB + local Excel into the current shape.

What this does, in order:

  1. **Backfill DB columns** for every existing row:
       - date_given      <- created_at when blank
       - source_link     <- rebuild from source_ref_id for Email rows
       - source_detail   <- "Google Chat" fallback for Chat rows that had it
                            blank (we can't reconstruct the DM/space partner;
                            this at least makes the Source column readable)

  2. **Reset sync state** on every task:
       - synced_to_sheets    = 0
       - sheet_row_all       = NULL
       - sheet_row_source    = NULL
     This forces a clean re-push of every row on the next forward sync,
     so the new 13-column / 4-tab layout is populated correctly without
     leaving orphan rows tied to old sheet positions.

  3. **Rebuild local Excel** (`tasks.xlsx`):
       - Renames legacy tabs to the current names
       - Inserts the missing columns to bring 10-col tabs onto 13-col HEADERS
       - **Wipes all data rows** (keeps only the header row in each tab)
     The next forward sync re-populates everything from the freshly-clean
     DB rows.

After this script runs, restart `python main.py`. On boot:
  - Sheets/Excel ensure_tabs() will create any missing managed tabs
    and confirm column headers
  - The forward sync worker will push every DB row in batches; each row
    lands in the correct tab per the routing (Email + Chat -> Gmail;
    Meeting -> In-Person Meetings; everything also goes to Master Task
    List)

Google Sheets caveat: this script can't reach Google Sheets without
OAuth. To avoid duplicate rows on the cloud side, manually delete every
data row (keep row 1 / headers) on each tab BEFORE restarting main.py.
The forward sync will then repopulate the Sheet from scratch. If you
skip that step, the Sheet just gets new rows appended on top of the old
ones — annoying but recoverable (delete duplicates manually later).

Run:
    python -m scripts.migrate_v2
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from database import get_db  # noqa: E402
from services.task_service import normalize_heading, _format_update_line  # noqa: E402
from sheets.client import HEADERS, TAB_ORDER  # noqa: E402
from sheets.excel_mirror import DEFAULT_PATH, ExcelMirror  # noqa: E402
from sheets.sync import _format_source_label  # noqa: E402
from transcription.lexicon import canonical_spoc  # noqa: E402
from utils.identifiers import clean_identifier  # noqa: E402
from utils.logger import get_logger  # noqa: E402

logger = get_logger("migrate_v2")

_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
_PHONE_RE = re.compile(r"(?:\+?\d[\s\-.]?){8,15}")


def _recover_contact(*sources):
    for s in sources:
        if not s:
            continue
        m = _EMAIL_RE.search(s)
        if m:
            return m.group(0)
    for s in sources:
        if not s:
            continue
        m = _PHONE_RE.search(s)
        if m:
            return m.group(0).strip()
    return None


def _email_link(message_id: str) -> str:
    return f"https://mail.google.com/mail/u/0/#inbox/{message_id}"


def step1_backfill_db() -> dict:
    """Backfill missing columns where we can recover the value."""
    db = get_db()
    rows = db.fetchall(
        """
        SELECT id, source_type, source_ref_id, source_detail, source_link,
               date_given, summary, spoc_contact, created_at
          FROM extracted_tasks
        """
    )

    fixed = {"date_given": 0, "source_link": 0, "source_detail": 0, "spoc_contact": 0}

    for r in rows:
        sets, params = [], []

        if not (r["date_given"] or "").strip() and r["created_at"]:
            sets.append("date_given = ?")
            params.append(r["created_at"])
            fixed["date_given"] += 1

        if not (r["spoc_contact"] or "").strip():
            recovered = _recover_contact(r["source_detail"], r["summary"])
            if recovered:
                sets.append("spoc_contact = ?")
                params.append(recovered)
                fixed["spoc_contact"] += 1

        if (
            (r["source_type"] or "").lower() == "email"
            and not (r["source_link"] or "").strip()
            and r["source_ref_id"]
        ):
            sets.append("source_link = ?")
            params.append(_email_link(r["source_ref_id"]))
            fixed["source_link"] += 1

        if (
            (r["source_type"] or "").lower() == "chat"
            and not (r["source_detail"] or "").strip()
        ):
            sets.append("source_detail = ?")
            params.append("Google Chat")
            fixed["source_detail"] += 1

        if not sets:
            continue
        params.append(r["id"])
        db.execute(
            f"UPDATE extracted_tasks SET {', '.join(sets)} WHERE id = ?",
            tuple(params),
        )

    logger.info(
        "Step 1 backfill: date_given=%d, source_link=%d, source_detail=%d, spoc_contact=%d",
        fixed["date_given"], fixed["source_link"], fixed["source_detail"], fixed["spoc_contact"],
    )
    return fixed


def step1a_sanitize_identifiers() -> int:
    """
    Walk every row and run sender_or_speaker / spoc_contact through
    `clean_identifier`. Strips opaque junk like `users/<id>`, `(unknown)`,
    UUID-like strings — anything that isn't a real name/email/phone.
    Returns the number of rows patched.
    """
    db = get_db()
    rows = db.fetchall(
        """
        SELECT id, sender_or_speaker, spoc_contact
          FROM extracted_tasks
        """
    )
    patched = 0
    for r in rows:
        # Identifier strip first (drop "users/12345" junk), then canonical
        # display collapse ("Anchit Tandon" -> "Anchit", "Aman Gupta" ->
        # "Aman", "Anchit (Self)" -> "Anchit").
        new_spoc = canonical_spoc(clean_identifier(r["sender_or_speaker"]))
        new_contact = clean_identifier(r["spoc_contact"])
        if (
            (new_spoc or None) == (r["sender_or_speaker"] or None)
            and (new_contact or None) == (r["spoc_contact"] or None)
        ):
            continue
        db.execute(
            """
            UPDATE extracted_tasks
               SET sender_or_speaker = ?,
                   spoc_contact      = ?
             WHERE id = ?
            """,
            (new_spoc, new_contact, int(r["id"])),
        )
        patched += 1
    logger.info(
        "Step 1a sanitize: %d row(s) had SPOC/contact normalized or stripped.",
        patched,
    )
    return patched


def step1b_dedup_existing_tasks() -> dict:
    """
    Walk every OPEN task. Group by (normalized_heading, lowercase SPOC).
    Within each group: keep the OLDEST row as the canonical task; for
    every other row in the group, append a structured update line to the
    canonical row's all_updates and DELETE the duplicate row.

    Also: for the canonical row, populate `normalized_heading` (so future
    inserts via TaskService.merge logic find the same group key) and
    rebuild its all_updates from every contributing row's metadata,
    chronologically ordered.

    Result: at most one open row per (heading, SPOC) pair.
    """
    db = get_db()
    rows = db.fetchall(
        """
        SELECT id, source_type, source_detail, task, task_description,
               sender_or_speaker, date_given, created_at, summary,
               all_updates, status
          FROM extracted_tasks
         WHERE status = 'open'
         ORDER BY COALESCE(date_given, created_at) ASC, id ASC
        """
    )

    # Build groups keyed on (normalized heading, lowercase SPOC).
    groups: dict[tuple[str, str], list] = {}
    for r in rows:
        key = (
            normalize_heading(r["task"]),
            (clean_identifier(r["sender_or_speaker"]) or "").lower(),
        )
        if not key[0]:
            continue
        groups.setdefault(key, []).append(r)

    canonical_kept = 0
    duplicates_merged = 0
    duplicates_deleted = 0

    for (norm, spoc_lc), group in groups.items():
        canonical = group[0]
        canonical_id = int(canonical["id"])
        canonical_kept += 1

        # Build the canonical row's All Updates from every group member.
        update_lines: list[str] = []
        for r in group:
            line = _format_update_line(
                when=r["date_given"] or r["created_at"],
                speaker=clean_identifier(r["sender_or_speaker"]),
                source_label=_format_source_label(
                    source_type=r["source_type"] or "",
                    source_detail=r["source_detail"] or "",
                ),
                note=(r["task_description"] or r["summary"] or "").strip(),
            )
            if line:
                update_lines.append(line)
        # Preserve any pre-existing all_updates on the canonical row.
        if canonical["all_updates"]:
            update_lines.insert(0, canonical["all_updates"].strip())
        # De-dup identical lines while preserving order.
        seen: set[str] = set()
        deduped: list[str] = []
        for ln in update_lines:
            if ln and ln not in seen:
                seen.add(ln)
                deduped.append(ln)
        new_updates = "\n".join(deduped) if deduped else None

        # Patch canonical row.
        db.execute(
            """
            UPDATE extracted_tasks
               SET normalized_heading = ?,
                   all_updates        = ?,
                   synced_to_sheets   = 0
             WHERE id = ?
            """,
            (norm, new_updates, canonical_id),
        )

        # Delete every other row in the group.
        if len(group) > 1:
            dup_ids = [int(r["id"]) for r in group[1:]]
            duplicates_merged += len(dup_ids)
            duplicates_deleted += len(dup_ids)
            qmarks = ",".join("?" * len(dup_ids))
            db.execute(
                f"DELETE FROM extracted_tasks WHERE id IN ({qmarks})",
                tuple(dup_ids),
            )

    logger.info(
        "Step 1b dedup: %d canonical task(s) kept, %d duplicate(s) merged + deleted.",
        canonical_kept, duplicates_deleted,
    )
    return {
        "canonical": canonical_kept,
        "merged": duplicates_merged,
    }


def step2_reset_sync_state() -> int:
    """Force every task to re-push to Sheets/Excel cleanly on next sync."""
    db = get_db()
    n = db.fetchone("SELECT COUNT(*) c FROM extracted_tasks WHERE synced_to_sheets = 1")["c"]
    db.execute(
        """
        UPDATE extracted_tasks
           SET synced_to_sheets = 0,
               sheet_row_all    = NULL,
               sheet_row_source = NULL
        """
    )
    logger.info("Step 2 reset: %d task(s) marked unsynced.", n)
    return n


def step3_rebuild_excel() -> bool:
    """
    Rebuild `tasks.xlsx` from scratch with the 4 canonical tabs and the
    13-column HEADERS. We deliberately do NOT try to migrate the existing
    file in place — that path turned out to be brittle (openpyxl can
    leave orphan worksheet parts behind, especially when the original
    file passed through several legacy schemas). Rebuilding atomically
    is simpler and bulletproof: the DB has every task, sync re-pushes
    them all on next boot.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    path = DEFAULT_PATH

    # Build fresh.
    wb = Workbook()
    # Replace the default "Sheet" with the first managed tab, then add the rest.
    default = wb.active
    default.title = TAB_ORDER[0]

    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="EAEAEA", end_color="EAEAEA", fill_type="solid"
    )
    header_align = Alignment(vertical="center")

    def _write_header(ws) -> None:
        for col, value in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=value)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"

    _write_header(default)
    for tab in TAB_ORDER[1:]:
        ws = wb.create_sheet(title=tab)
        _write_header(ws)

    # Be paranoid — assert exactly the right tab set before saving.
    assert list(wb.sheetnames) == list(TAB_ORDER), (
        f"workbook has unexpected tabs: {wb.sheetnames}"
    )

    # Write to a sibling file then atomically replace, so a half-written
    # workbook never appears at the canonical path.
    tmp = path.with_suffix(path.suffix + ".tmp")
    try:
        wb.save(tmp)
    except PermissionError:
        logger.warning(
            "Could not save %s — is it open in Excel? Close and re-run.", tmp
        )
        return False

    try:
        tmp.replace(path)
    except PermissionError:
        logger.warning(
            "Could not replace %s with rebuilt file — close it in Excel and re-run.", path
        )
        return False

    logger.info(
        "Excel rebuilt from scratch: %s -> %d tabs, %d cols, no data rows.",
        path, len(TAB_ORDER), len(HEADERS),
    )

    # Reset the singleton flag so any subsequent ExcelMirror call sees the
    # new file rather than its cached "initialised" view.
    ExcelMirror._initialised = False  # type: ignore[attr-defined]
    return True


def main() -> int:
    print("=== Personal AI OS — v2 schema migration ===")
    s1 = step1_backfill_db()
    sa = step1a_sanitize_identifiers()
    sd = step1b_dedup_existing_tasks()
    s2 = step2_reset_sync_state()
    ok = step3_rebuild_excel()

    print()
    print("Summary:")
    print(f"  Step 1  backfill : date_given={s1['date_given']}, source_link={s1['source_link']}, "
          f"source_detail={s1['source_detail']}, spoc_contact={s1['spoc_contact']}")
    print(f"  Step 1a sanitize : {sa} row(s) had placeholder SPOC/contact stripped")
    print(f"  Step 1b dedup    : {sd['canonical']} canonical task(s) kept, "
          f"{sd['merged']} duplicate(s) merged into All Updates")
    print(f"  Step 2  reset    : {s2} task(s) marked unsynced")
    print(f"  Step 3  Excel    : {'rebuilt' if ok else 'SKIPPED (see log)'}")
    print()
    print("Next steps:")
    print("  1. (Optional but recommended) In Google Sheets, manually delete every")
    print("     DATA row in each tab (keep the header row). This avoids duplicates.")
    print("  2. Restart:  python main.py")
    print("     The forward sync worker will repopulate everything from the DB.")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
