"""
Local Excel mirror of the Google Sheets surface.

Writes the same task rows that go to Google Sheets into a workbook at
the repo root (`tasks.xlsx`), preserving the same 4-tab layout and 13
columns (see HEADERS in sheets/client.py). The file lives in git so the task list is review-able from
GitHub directly, and a clone of the repo always has the latest
snapshot baked in.

Failure modes (and how they're handled):
  - File is open in Excel (Windows file lock) -> log warning, skip the
    write, keep going. The next successful flush will catch it up.
  - File is corrupt / unreadable -> rebuild from scratch.

The mirror is intentionally append-only. Status updates that come in
from the sheet edit-back loop (a future feature) will need a separate
update method.
"""
from __future__ import annotations

import threading
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from config import settings
from utils.logger import get_logger

from .client import HEADERS, LEGACY_SCHEMAS, LEGACY_TAB_RENAMES, TAB_ORDER

logger = get_logger(__name__)

DEFAULT_PATH = settings.project_root / "tasks.xlsx"

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
_HEADER_ALIGN = Alignment(vertical="center")


def _column_widths() -> dict[str, int]:
    """Reasonable default widths so the Excel doesn't open as a postage stamp."""
    return {
        "A": 36,  # Task Heading
        "B": 60,  # Task Description
        "C": 12,  # Status
        "D": 28,  # Source (Email | Chat | Meeting + detail)
        "E": 32,  # Source Link
        "F": 22,  # Task Given On
        "G": 50,  # Why We're Doing This
        "H": 18,  # Growth Pillar
        "I": 22,  # SPOC
        "J": 28,  # SPOC Contact
        "K": 12,  # Priority
        "L": 18,  # Task Deadline
        "M": 60,  # All Updates (chronological log)
        "N": 30,  # Remarks
    }


class ExcelMirror:
    """
    Thread-safe append-only writer for `tasks.xlsx`.

    Concurrency model: a single in-process lock serialises every read +
    mutation cycle. The Sheets sync worker is the only producer in
    practice, so this is plenty.
    """

    def __init__(self, path: Optional[Path] = None) -> None:
        self._path: Path = (path or DEFAULT_PATH).resolve()
        self._lock = threading.Lock()
        self._initialised = False

    # --- public API ----------------------------------------------------------

    def append_rows(self, tab: str, rows: list[list[str]]) -> None:
        """Append rows to the given tab in the workbook. Idempotent on lock errors."""
        if not rows:
            return
        with self._lock:
            self._ensure_workbook()
            try:
                wb = load_workbook(self._path)
            except (FileNotFoundError, InvalidFileException):
                wb = self._build_empty_workbook()

            if tab not in wb.sheetnames:
                self._add_tab(wb, tab)
            ws = wb[tab]
            self._heal_legacy_schema(ws)
            for row in rows:
                ws.append(row)

            try:
                wb.save(self._path)
                logger.info(
                    "Excel mirror: appended %d row(s) to %r in %s",
                    len(rows),
                    tab,
                    self._path.name,
                )
            except PermissionError:
                # Most common cause: the file is open in Excel and Windows
                # has a write lock on it. Skip silently — the Google Sheet
                # is the source of truth and the next flush will catch us up.
                logger.warning(
                    "Excel mirror: %s is locked (probably open in Excel). "
                    "Close the file to resume mirroring; Google Sheets keeps working.",
                    self._path,
                )

    # --- bootstrap -----------------------------------------------------------

    def _ensure_workbook(self) -> None:
        if self._initialised and self._path.exists():
            return
        if not self._path.exists():
            wb = self._build_empty_workbook()
            try:
                wb.save(self._path)
                logger.info("Created Excel mirror at %s", self._path)
            except PermissionError:
                logger.warning(
                    "Could not create Excel mirror at %s (locked).", self._path
                )
                return
        else:
            # Existing file — make sure it has all expected tabs in order.
            try:
                wb = load_workbook(self._path)
            except InvalidFileException:
                logger.warning(
                    "Existing %s is unreadable; rebuilding from scratch.",
                    self._path.name,
                )
                wb = self._build_empty_workbook()
                try:
                    wb.save(self._path)
                except PermissionError:
                    return

            mutated = False

            # Rename legacy worksheet titles in place so existing rows are
            # preserved across the rename. Skip if the new name already
            # exists (manual migration / prior boot already did it).
            for old, new in LEGACY_TAB_RENAMES.items():
                if old == new:
                    continue
                if old in wb.sheetnames and new not in wb.sheetnames:
                    logger.info(
                        "Excel mirror: renaming legacy tab %r -> %r", old, new
                    )
                    wb[old].title = new
                    mutated = True

            # Add any missing managed tabs.
            for tab in TAB_ORDER:
                if tab not in wb.sheetnames:
                    self._add_tab(wb, tab)
                    mutated = True

            # Drop the default empty "Sheet" tab Excel injects on creation
            # if it's still there alongside our managed tabs.
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > len(TAB_ORDER):
                del wb["Sheet"]
                mutated = True

            if mutated:
                try:
                    wb.save(self._path)
                except PermissionError:
                    pass

        self._initialised = True

    def _build_empty_workbook(self) -> Workbook:
        wb = Workbook()
        # Workbook() creates a default "Sheet". Replace it with the first
        # managed tab and add the rest.
        default = wb.active
        default.title = TAB_ORDER[0]
        self._write_header(default)
        for tab in TAB_ORDER[1:]:
            ws = wb.create_sheet(title=tab)
            self._write_header(ws)
        # Reorder to canonical order (create_sheet appends, which is fine,
        # but be explicit for safety).
        wb._sheets = [wb[t] for t in TAB_ORDER]  # type: ignore[attr-defined]
        return wb

    def _add_tab(self, wb: Workbook, tab: str) -> None:
        ws = wb.create_sheet(title=tab)
        self._write_header(ws)
        # Sort all managed tabs to the front in canonical order.
        managed_in_order = [wb[t] for t in TAB_ORDER if t in wb.sheetnames]
        leftovers = [s for s in wb.worksheets if s not in managed_in_order]
        wb._sheets = managed_in_order + leftovers  # type: ignore[attr-defined]

    def _heal_legacy_schema(self, ws) -> None:
        """
        Bring an existing worksheet onto the current HEADERS layout.

        Drift cases handled:
          1. Any layout listed in LEGACY_SCHEMAS — blank columns are
             inserted at the recorded indices to realign existing data
             with the new HEADERS, then the header row is rewritten.
          2. Header is the correct prefix but has trailing junk cells:
             truncate to len(HEADERS).
        """
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

        # Case 1: known legacy layout — shift data right and rewrite header.
        for legacy_header, insert_at in LEGACY_SCHEMAS:
            if header[: len(legacy_header)] == legacy_header:
                logger.info(
                    "Excel mirror: tab %r is on legacy %d-col schema; "
                    "inserting %d blank column(s) to migrate.",
                    ws.title, len(legacy_header), len(insert_at),
                )
                # openpyxl's insert_cols is 1-indexed; LEGACY_SCHEMAS is 0-indexed.
                # Apply ascending so each subsequent index is correct in the
                # post-insert coordinate space.
                for idx in sorted(insert_at):
                    ws.insert_cols(idx + 1)
                self._write_header(ws)
                return

        # Case 2: header is correct prefix but has trailing extras.
        if header[: len(HEADERS)] == HEADERS and ws.max_column > len(HEADERS):
            logger.info(
                "Excel mirror: tab %r has trailing columns past %d; trimming.",
                ws.title,
                len(HEADERS),
            )
            for extra_col in range(ws.max_column, len(HEADERS), -1):
                ws.cell(row=1, column=extra_col).value = None

    def _write_header(self, ws) -> None:
        for col, value in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=value)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
        for letter, width in _column_widths().items():
            ws.column_dimensions[letter].width = width
        ws.freeze_panes = "A2"


_singleton: Optional[ExcelMirror] = None
_singleton_lock = threading.Lock()


def get_excel_mirror() -> ExcelMirror:
    global _singleton
    if _singleton is None:
        with _singleton_lock:
            if _singleton is None:
                _singleton = ExcelMirror()
    return _singleton
