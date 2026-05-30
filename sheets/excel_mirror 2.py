"""
Local Excel mirror of the Google Sheets surface.

Writes the same task rows that go to Google Sheets into `tasks.xlsx`,
preserving the same 3-tab layout and hidden metadata columns. Rows are
upserted by hidden task id so reminder replies and checklist edits are
mirrored cleanly.
"""
from __future__ import annotations

import threading
from pathlib import Path
from typing import Optional
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from config import settings
from utils.logger import get_logger

from .client import HEADERS, LEGACY_SCHEMAS, LEGACY_TAB_RENAMES, TAB_ORDER

logger = get_logger(__name__)

# Exceptions that mean "the .xlsx on disk is unreadable — rebuild from
# scratch rather than crashing." openpyxl raises InvalidFileException
# for the obvious non-zip case, but BadZipFile (from the stdlib zipfile
# module) and KeyError can also fire when the file is half-written or
# the central directory is corrupt. Catching all three keeps a single
# corrupt file from cascading into a main.py crash.
_XLSX_CORRUPT_ERRORS: tuple[type[BaseException], ...] = (
    InvalidFileException,
    BadZipFile,
    KeyError,            # openpyxl raises this when zip entries are missing
    OSError,             # "Bad magic number" sometimes lands here
)

DEFAULT_PATH = settings.project_root / "tasks.xlsx"
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
_HEADER_ALIGN = Alignment(vertical="center")


def _column_widths() -> dict[str, int]:
    return {
        "A": 10,
        "B": 36,
        "C": 60,
        "D": 22,
        "E": 12,
        "F": 28,
        "G": 32,
        "H": 50,
        "I": 18,
        "J": 22,
        "K": 28,
        "L": 12,
        "M": 18,
        "N": 60,
        "O": 30,
    }


class ExcelMirror:
    def __init__(self, path: Optional[Path] = None) -> None:
        self._path: Path = (path or DEFAULT_PATH).resolve()
        self._lock = threading.Lock()
        self._initialised = False

    def upsert_task_row(self, tab: str, row: list[object]) -> None:
        if not row:
            return
        with self._lock:
            self._ensure_workbook()
            try:
                wb = load_workbook(self._path)
            except FileNotFoundError:
                wb = self._build_empty_workbook()
            except _XLSX_CORRUPT_ERRORS as exc:
                logger.warning(
                    "Excel mirror: %s is corrupt (%s); rebuilding from scratch.",
                    self._path.name, type(exc).__name__,
                )
                # Park the corrupt file as a .bad sibling so the user
                # can inspect it later if they care; don't lose it
                # silently.
                try:
                    bad = self._path.with_suffix(self._path.suffix + ".bad")
                    self._path.replace(bad)
                except Exception:
                    pass
                wb = self._build_empty_workbook()

            if tab not in wb.sheetnames:
                self._add_tab(wb, tab)
            ws = wb[tab]
            self._heal_legacy_schema(ws)

            task_id = str(row[-1]).strip()
            target_row = None
            if task_id:
                for idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=idx, column=len(HEADERS)).value
                    if str(cell or "").strip() == task_id:
                        target_row = idx
                        break

            if target_row is None:
                ws.append(row)
            else:
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=target_row, column=col_idx, value=value)

            try:
                wb.save(self._path)
            except PermissionError:
                logger.warning(
                    "Excel mirror: %s is locked. Close it to resume mirroring.",
                    self._path,
                )

    def _ensure_workbook(self) -> None:
        if self._initialised and self._path.exists():
            return
        if not self._path.exists():
            wb = self._build_empty_workbook()
            try:
                wb.save(self._path)
            except PermissionError:
                return
        else:
            try:
                wb = load_workbook(self._path)
            except InvalidFileException:
                wb = self._build_empty_workbook()
                try:
                    wb.save(self._path)
                except PermissionError:
                    return

            mutated = False
            for old, new in LEGACY_TAB_RENAMES.items():
                if old == new:
                    continue
                if old in wb.sheetnames and new not in wb.sheetnames:
                    wb[old].title = new
                    mutated = True

            for tab in TAB_ORDER:
                if tab not in wb.sheetnames:
                    self._add_tab(wb, tab)
                    mutated = True

            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > len(TAB_ORDER):
                del wb["Sheet"]
                mutated = True

            if mutated:
                try:
                    wb.save(self._path)
                except PermissionError:
                    pass

        self._initialised = True

    def _build_empty_workbook(self) -> Workbook:
        wb = Workbook()
        default = wb.active
        default.title = TAB_ORDER[0]
        self._write_header(default)
        for tab in TAB_ORDER[1:]:
            ws = wb.create_sheet(title=tab)
            self._write_header(ws)
        wb._sheets = [wb[t] for t in TAB_ORDER]  # type: ignore[attr-defined]
        return wb

    def _add_tab(self, wb: Workbook, tab: str) -> None:
        ws = wb.create_sheet(title=tab)
        self._write_header(ws)
        managed_in_order = [wb[t] for t in TAB_ORDER if t in wb.sheetnames]
        leftovers = [s for s in wb.worksheets if s not in managed_in_order]
        wb._sheets = managed_in_order + leftovers  # type: ignore[attr-defined]

    def _heal_legacy_schema(self, ws) -> None:
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for legacy_header, insert_at in LEGACY_SCHEMAS:
            if header[: len(legacy_header)] == legacy_header:
                for idx in sorted(insert_at):
                    ws.insert_cols(idx + 1)
                self._write_header(ws)
                return

        if header[: len(HEADERS)] == HEADERS and ws.max_column > len(HEADERS):
            for extra_col in range(ws.max_column, len(HEADERS), -1):
                ws.cell(row=1, column=extra_col).value = None

    def _write_header(self, ws) -> None:
        for col, value in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=value)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
        for letter, width in _column_widths().items():
            ws.column_dimensions[letter].width = width
        ws.column_dimensions["P"].hidden = True
        ws.column_dimensions["Q"].hidden = True
        ws.freeze_panes = "F2"


_singleton: Optional[ExcelMirror] = None
_singleton_lock = threading.Lock()


def get_excel_mirror() -> ExcelMirror:
    global _singleton
    if _singleton is None:
        with _singleton_lock:
            if _singleton is None:
                _singleton = ExcelMirror()
    return _singleton
